from pathlib import Path
import openpyxl
import numpy
import json
import module_xlsx

def get_xlsx_colname(filepath, row_num):
    """
    엑셀(xlsx) 파일의 컬럼 이름을 얻어옵니다.

    Args:
        filepath(str): xlsx 파일 경로
        row_num: 컬럼이 위치한 행 번호

    Return:
        <class 'list'>
    """
    load_wb = openpyxl.load_workbook(filepath)
    load_ws = load_wb.active  # 첫 번째 시트

    # 컬럼명 찍기
    column_data = load_ws[row_num:row_num]
    
    title_list = []
    for item in column_data:
        title_list.append(item.value)
    
    print("# 컬럼 이름")
    print(f"{title_list}\n")
    
    return title_list


def get_xlsx_data(filepath, row_num):
    """
    엑셀(xlsx) 파일에서 컬럼 이름을 제외한 데이터를 얻어옵니다.

    Args:
        filepath(str): xlsx 파일 경로
        row_num: 컬럼이 위치한 행 번호

    Return:
        <class 'list'>
    """
    load_wb = openpyxl.load_workbook(filepath)
    load_ws = load_wb.active  # 첫 번째 시트

    # 데이터 추출
    data = load_ws[row_num+1:load_ws.max_row]
    lenM = numpy.ndim(data)

    if lenM == 1:  # 데이터 행이 1개인 경우 (1차원 데이터 처리)
        data_lists = []
        data_list = []

        for cell in data:
            data_list.append(cell.value)

        if data_list:
            data_lists.append(data_list)
        
        print(f"{filepath} 파일에서 {len(data_lists)}개의 행을 얻어옴")
        return data_lists

    elif lenM == 2:  # 데이터 행이 2개 이상인 경우 (2차원 데이터 처리)
        data_lists = []

        for row in data:
            data_list = []

            for cell in row:
                data_list.append(cell.value)
            
            if data_list:
                data_lists.append(data_list)

        print(f"{filepath} 파일에서 {len(data_lists)}개의 행을 얻어옴")
        return data_lists


def merge_xlsx():
    """
    특정 폴더에 존재하는 모든 엑셀(XLSX) 파일들의 수합을 시도합니다.
    특정 폴더는 config.json 파일에 명시되어 있습니다. (기본: data 폴더)

    Args:
        None

    Return:
        None
    """
    # 설정 파일을 불러온 뒤 xlsx 파일경로 정보를 담을 Path 객체 생성
    with open('config.json') as config_file:
        config = json.load(config_file)
    path = Path(config['folder_name'])

    # 어느 폴더에 어떤 파일로 수합 저장하는지 알림
    print(f"\n이 프로그램은 {config['folder_name']} 폴더에 있는 모든 엑셀(XLSX) 파일의 내용을")
    print(f"{config['save_name_xlsx']} 파일에 수합 저장합니다.\n")

    # 최종 결과를 수합할 xlsx 생성
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.create_sheet(config["sheet_name"], 0)

    # 몇 번째 행이 이름 행인지 입력받고 찍기
    row_num = int(input("몇 번째 행이 컬럼 이름? (숫자로만 입력) => "))
    for xlsx in path.glob("*.xlsx"):
        new_ws.append(module_xlsx.get_xlsx_colname(xlsx, row_num))
        break

    # 컬럼 이름 행이 아닌 데이터 수집해서 시트에 합치기
    print("# 데이터 읽어오기")
    for xlsx in path.glob("*.xlsx"):
        rows = module_xlsx.get_xlsx_data(xlsx, row_num)
        for row in rows:
            new_ws.append(row)

    # 최종결과 저장
    new_wb.save(config["save_name_xlsx"])
    print("\n# 수합된 데이터 저장")
    print(f"{config['save_name_xlsx']} 파일에 저장 완료!")
    exit(0)
